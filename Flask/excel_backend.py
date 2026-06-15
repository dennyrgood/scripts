from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from datetime import datetime, date
import os
import shutil
import logging

app = Flask(__name__)
CORS(app)

# Logging setup
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration
EXCEL_FILE = os.path.expanduser(r"D:\OneDrive\DropBoxReplacement\MathesDropBox\00  Top Drawer\Movies Shows - to add.xlsx")
#"D:/OneDrive/MS/MoviesShows.xlsx")

# ============ CORE EXCEL FUNCTIONS (from your Tkinter script) ============

def create_backup(original_file):
    """Create timestamped backup inside a dedicated 'Backup' directory."""
    # 1. Force absolute path to handle OneDrive/Windows path quirks
    abs_original_path = os.path.abspath(original_file)
    
    if not os.path.exists(abs_original_path):
        logger.error(f"Backup Source Missing: {abs_original_path}")
        return False, f"File not found: {abs_original_path}"
    
    try:
        # 2. Get the directory of the file
        file_dir = os.path.dirname(abs_original_path)
        
        # 3. Create 'Backup' folder relative to the file's actual location
        backup_dir = os.path.join(file_dir, "Backup")
        os.makedirs(backup_dir, exist_ok=True)

        # 4. Extract filename and prepare timestamp
        filename = os.path.basename(abs_original_path)
        name, ext = os.path.splitext(filename)
        timestamp = datetime.now().strftime(" - Backup %Y %m %d %H %M %S")
        
        # 5. Join the Backup folder with the new filename
        backup_file = os.path.join(backup_dir, f"{name}{timestamp}{ext}")
        
        # 6. Copy the file
        shutil.copy2(abs_original_path, backup_file)
        
        logger.info(f"Successfully backed up to: {backup_file}")
        return True, backup_file
        
    except Exception as e:
        logger.error(f"Backup failed. Reason: {str(e)}")
        return False, f"Backup failed: {str(e)}"

def copy_cell_properties(source_cell, target_cell, copy_number_format=True):
    """Copy formatting from source to target cell."""
    if source_cell.has_style:
        target_cell.style = source_cell.style
    if copy_number_format and source_cell.number_format:
        target_cell.number_format = source_cell.number_format

def copy_formulas(ws, source_row, target_row):
    """Copy formulas from B,C,D,E with updated row references."""
    formula_cols = [2, 3, 4, 5]
    old_row_str = str(source_row)
    new_row_str = str(target_row)
    columns_to_update = ['A', 'B']
    replacement_targets = []
    
    for col in columns_to_update:
        replacement_targets.append((f"{col}{old_row_str}", f"{col}{new_row_str}"))
        replacement_targets.append((f"{col.lower()}{old_row_str}", f"{col.lower()}{new_row_str}"))
    
    for col_idx in formula_cols:
        source_cell = ws.cell(row=source_row, column=col_idx)
        target_cell = ws.cell(row=target_row, column=col_idx)
        
        if source_cell.data_type == 'f' and source_cell.value:
            formula_text = source_cell.value
            updated_formula = formula_text
            for old_ref, new_ref in replacement_targets:
                updated_formula = updated_formula.replace(old_ref, new_ref)
            target_cell.value = updated_formula
        else:
            target_cell.value = source_cell.value

def parse_coordinate(coord_str):
    """Parse Excel coordinate (e.g., 'A10') into (col, row)."""
    col = ""
    row_str = ""
    for char in coord_str:
        if char.isalpha():
            col += char.upper()
        elif char.isdigit():
            row_str += char
    if not col or not row_str:
        raise ValueError(f"Invalid coordinate: {coord_str}")
    return (col, int(row_str))

def insert_row_to_excel(target_file, new_row_data):
    """Add new row to Excel with formatting and formula preservation."""
    try:
        wb = load_workbook(target_file, data_only=False)
        ws = wb.active
        
        table_name = None
        table_range = None
        table_style = None
        max_row = ws.max_row
        row_start_num = 1
        
        if ws.tables:
            table_name = list(ws.tables.keys())[0]
            table = ws.tables[table_name]
            table_range = table.ref
            if hasattr(table, 'tableStyleInfo'):
                table_style = table.tableStyleInfo
            start_coord_str, end_coord_str = table_range.split(':')
            start_col_letter, row_start_num = parse_coordinate(start_coord_str)
            end_col_letter, max_row = parse_coordinate(end_coord_str)
        else:
            start_col_letter, end_col_letter = 'A', 'M'
        
        last_existing_row = max_row
        if last_existing_row <= 1:
            new_row_number = 2
            formula_source_row = 1
        else:
            new_row_number = last_existing_row + 1
            formula_source_row = last_existing_row
        
        if formula_source_row >= 1:
            for col_idx in range(1, 14):
                source_cell = ws.cell(row=formula_source_row, column=col_idx)
                target_cell = ws.cell(row=new_row_number, column=col_idx)
                is_column_j = (col_idx == 10)
                copy_cell_properties(source_cell, target_cell, copy_number_format=not is_column_j)
            
            if formula_source_row > 1:
                copy_formulas(ws, formula_source_row, new_row_number)
        
        ws.cell(row=new_row_number, column=1, value=new_row_data[0])
        for i, value in enumerate(new_row_data[1:]):
            col_idx = 6 + i
            target_cell = ws.cell(row=new_row_number, column=col_idx, value=value)
            if col_idx == 10 and isinstance(value, date):
                target_cell.number_format = 'dd-mmm-yyyy'
        
        if table_name and table_range:
            new_max_row = new_row_number
            new_ref = f"{start_col_letter}{row_start_num}:{end_col_letter}{new_max_row}"
            del ws.tables[table_name]
            new_table = Table(displayName=table_name, ref=new_ref)
            if table_style:
                new_table.tableStyleInfo = table_style
            ws.add_table(new_table)
        
        wb.save(target_file)
        return True, f"Row {new_row_number} added successfully"
    
    except Exception as e:
        logger.error(f"Excel processing error: {str(e)}")
        return False, f"Error: {str(e)}"

def parse_date(date_str):
    """Parse flexible date formats."""
    if not date_str:
        return None
    
    formats = [
        '%d-%b-%Y', '%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d',
        '%d/%m/%Y', '%d/%m/%y'
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    
    return None

# ============ FLASK ROUTES ============

@app.route('/api/submit', methods=['POST'])
def submit_data():
    """Handle form submission from web form."""
    try:
        data = request.json
        
        # Validate mandatory fields
        code_str = data.get('code', '').strip()
        title = data.get('title', '').strip()
        
        if not code_str or not title:
            return jsonify({"status": "error", "message": "Code and Title are required"}), 400
        
        # Validate code is numeric
        try:
            code = float(code_str)
        except ValueError:
            return jsonify({"status": "error", "message": "Code must be a number"}), 400
        
        # Parse optional fields
        col_g = data.get('col_g', '').strip()
        col_h = data.get('col_h', '').strip()
        col_i = data.get('col_i', 'Download').strip()
        col_j_str = data.get('col_j', '').strip()
        col_k = data.get('col_k', '').strip()
        col_l = data.get('col_l', '').strip()
        col_m = data.get('col_m', '').strip()
        
        # Parse date
        col_j = parse_date(col_j_str) if col_j_str else None
        if col_j_str and not col_j:
            return jsonify({"status": "error", "message": f"Invalid date format: {col_j_str}"}), 400
        
        # Create backup
        backup_ok, backup_msg = create_backup(EXCEL_FILE)
        if not backup_ok:
            logger.warning(f"Backup warning: {backup_msg}")
        
        # Insert row
        new_row_data = [code, title, col_g, col_h, col_i, col_j, col_k, col_l, col_m]
        success, message = insert_row_to_excel(EXCEL_FILE, new_row_data)
        
        if success:
            logger.info(f"Data inserted: {title}")
            return jsonify({"status": "success", "message": message}), 200
        else:
            return jsonify({"status": "error", "message": message}), 500
    
    except Exception as e:
        logger.error(f"API error: {str(e)}")
        return jsonify({"status": "error", "message": f"Server error: {str(e)}"}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Simple health check endpoint."""
    return jsonify({"status": "ok"}), 200

if __name__ == '__main__':
    # Check dependencies
    try:
        import flask_cors
    except ImportError:
        print("Installing flask-cors...")
        os.system("pip install flask-cors")
    
    logger.info(f"Starting Flask server")
    logger.info(f"Excel file: {EXCEL_FILE}")
    logger.info(f"Listening on 0.0.0.0:5000")
    
    app.run(host='0.0.0.0', port=5000, debug=False)